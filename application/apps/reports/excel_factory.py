from typing import Optional

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from apps.commons.utils.data_types.dict import DictUtils
from apps.commons.utils.data_types.list import ListUtils


class ExcelFactory:
    """Формирование excel файлов, их заполнение и выдача в виде response"""

    @staticmethod
    def get_row_and_col_for_merge(coordinates) -> Optional[list]:
        """Получение индекса строки и столбца для объединения ячеек"""
        try:
            spl = coordinates.split(':::')
            return [spl[0], spl[1]]
        except BaseException:
            return None

    @staticmethod
    def _get_cell_by_str_and_col(sheet, row: int, col: int):
        """Получение ячейки по заданным координатам и листу"""
        try:
            return sheet.cell(row=row, column=col)
        except BaseException:
            return sheet.cell(row=1, column=1)

    @staticmethod
    def _set_font_bold(cell):
        """Установка в ячейке жирный шрифт текста"""
        try:
            cell.font = Font(bold=True)
        except BaseException:
            pass

    @staticmethod
    def _set_font_color(cell, rgb, is_bold: bool):
        """Установка цвета текста ячейки (и жирного шрифта в случае получения True)"""
        try:
            if is_bold:
                cell.font = Font(bold=True, color=rgb)
            else:
                cell.font = Font(color=rgb)
        except BaseException:
            pass

    @staticmethod
    def _set_cell_alignment_center(cell):
        """Установить выравнивание текста в ячейке по центру по вертикали и горизонтали"""
        try:
            cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        except BaseException:
            pass

    @staticmethod
    def _set_sheet_title(sheet, title: str):
        """Установка названия листа книги"""
        try:
            sheet.title = title
        except BaseException:
            pass

    @staticmethod
    def _get_column_letter(column: int):
        """Получение буквенного названия столбца листа книги по его номеру"""
        return get_column_letter(column)

    def _merge_cells(self, sheet, row_start: int, col_start: int, row_finish: int, col_finish: int):
        """Объединение ячеек по их полученным индексам на листе книги"""
        try:
            str_col_start = self._get_column_letter(col_start)
            str_col_finish = self._get_column_letter(col_finish)
            sheet.merge_cells(f'{str_col_start}{row_start}:{str_col_finish}{row_finish}')
        except BaseException:
            pass

    @staticmethod
    def _set_row_height(sheet, row: int, height: int):
        """Установка высоты строки"""
        try:
            sheet.row_dimensions[row].height = height
        except BaseException:
            pass

    def _set_cell_width(self, sheet, col: int, value: int):
        """Установка ширины ячейки"""
        letter = self._get_column_letter(col)
        sheet.column_dimensions[letter].width = value

    @staticmethod
    def _set_cell_border(cell):
        """Отрисовка границ ячейки"""
        thin = Side(border_style="thin", color="000000")
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    def _set_cell_value(self, cell, value):
        """Установка значения ячейки"""
        try:
            cell.value = value
            self._set_cell_border(cell)
        except BaseException:
            pass

    @staticmethod
    def _set_cell_money_format(cell):
        """Установка денежного формата ячейки"""
        try:
            cell.number_format = '#,##0.00 [$₽-419]'
        except BaseException:
            pass

    @staticmethod
    def _fill_cell(cell, rgb):
        """Заливка ячейки полученным цветом"""
        cell.fill = PatternFill('solid', fgColor=rgb)

    def _fill_sheet(self, sheet, datacell: list):
        """Заполнение листа книги на основе полученной data"""
        du = DictUtils()
        if ListUtils.is_list_empty(datacell):
            return None
        for data in datacell:
            if isinstance(data, dict):
                if du.exist_key_in_dict('row', data) and du.exist_key_in_dict('col', data):
                    try:
                        cell = self._get_cell_by_str_and_col(
                            sheet,
                            du.get_int_value_in_dict_by_key('row', data),
                            du.get_int_value_in_dict_by_key('col', data)
                        )
                        if du.exist_key_in_dict('value', data):
                            self._set_cell_value(cell, du.get_value_in_dict_by_key('value', data))
                        if du.exist_key_in_dict('bold', data) and du.get_value_in_dict_by_key('bold', data):
                            self._set_font_bold(cell)
                        if du.exist_key_in_dict('money', data) and du.get_value_in_dict_by_key('money', data):
                            self._set_cell_money_format(cell)
                        if du.exist_key_in_dict('align_center', data) and \
                                du.get_value_in_dict_by_key('align_center', data):
                            self._set_cell_alignment_center(cell)
                        if du.exist_key_in_dict('width', data):
                            self._set_cell_width(
                                sheet,
                                du.get_int_value_in_dict_by_key('col', data),
                                du.get_int_value_in_dict_by_key('width', data),
                            )
                        if du.exist_key_in_dict('height', data):
                            self._set_row_height(
                                sheet,
                                du.get_int_value_in_dict_by_key('row', data),
                                du.get_int_value_in_dict_by_key('height', data),
                            )
                        if du.exist_key_in_dict('fill', data):
                            self._fill_cell(
                                cell,
                                du.get_str_value_in_dict_by_key('fill', data)
                            )
                        if du.exist_key_in_dict('color', data):
                            if du.exist_key_in_dict('bold', data) and du.get_value_in_dict_by_key('bold', data):
                                self._set_font_color(
                                    cell,
                                    du.get_str_value_in_dict_by_key('color', data),
                                    True
                                )
                            else:
                                self._set_font_color(
                                    cell,
                                    du.get_str_value_in_dict_by_key('color', data),
                                    False
                                )
                        if du.exist_key_in_dict('merge', data):
                            merge_coordinates = self.get_row_and_col_for_merge(
                                du.get_str_value_in_dict_by_key('merge', data)
                            )
                            if merge_coordinates is not None:
                                try:
                                    self._merge_cells(
                                        sheet,
                                        du.get_int_value_in_dict_by_key('row', data),
                                        du.get_int_value_in_dict_by_key('col', data),
                                        int(merge_coordinates[0]),
                                        int(merge_coordinates[1]),
                                    )
                                except BaseException:
                                    pass
                    except BaseException:
                        pass

    def _generate_workbook(self, data: list):
        """Генерация книги excel на основе полученной даты"""
        wb = Workbook()
        du = DictUtils()
        if ListUtils.is_list_empty(data):
            return wb
        for datasheet in data:
            if isinstance(datasheet, dict):
                if du.exist_key_in_dict('title', datasheet):
                    val = du.get_str_value_in_dict_by_key('title', datasheet)
                    if val is not None and len(val) > 0:
                        ws = wb.create_sheet(du.get_str_value_in_dict_by_key('title', datasheet), -1)
                    else:
                        ws = wb.create_sheet('Новый лист', -1)
                else:
                    ws = wb.create_sheet('Новый лист', -1)
                if du.exist_key_in_dict('sheet_data', datasheet):
                    self._fill_sheet(ws, du.get_list_in_dict_by_key('sheet_data', datasheet))
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        return wb

    def download_excel(self, filename: str, data: list) -> HttpResponse:
        """Формирование и скачивание excel файла на основе полученной даты и имени файла"""
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb = self._generate_workbook(data)
        wb.save(response)
        return response
